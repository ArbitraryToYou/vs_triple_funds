#基金管理人 url=https://danjuanfunds.com/djapi/fund/manager/history?fd_code=003984
#主要信息   url=https://danjuanfunds.com/djapi/fund/003984
#收益信息   url=https://danjuanfunds.com/djapi/fund/derived/003984
#十大重仓等 url=https://danjuanfunds.com/djapi/fund/detail/003984
#曲线       url=https://danjuanfunds.com/djapi/fund/nav-growth/003984?day=360

from win32com.client import Dispatch
import requests
import openpyxl
import json
import time
from datetime import datetime as dt
from datetime import timedelta as dd
import os

class Fund:
    def __init__(self, code, day):
        self.code = code
        self.day = day
        self.headers = {
            'Cookie': 'device_id=web_SJedqKpesv; gr_user_id=1d04b762-0a3b-41c4-b8b8-004fba79b385; accesstoken=240010000d68ddef5d0a921adff3a9ca3b71caf188ee68894; u=161474641; uid=161474641; refreshtoken=2400100001755f51276cba8a0a5b949dff63d230d9bada5c7; acw_tc=2760778316125744776097384e603ef6fb50648ef4438b3511b678a989a199; xq_a_token=244708005bc946786b8ba870b872c392f4f1e35e; Hm_lvt_d8a99640d3ba3fdec41370651ce9b2ac=1612351930,1612574478; channel=1300100141; Hm_lpvt_d8a99640d3ba3fdec41370651ce9b2ac=1612574978; timestamp=1612576174413',
            'Referer': 'https://danjuanfunds.com/funding/002876?channel=1300100141',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36'
        }
        self.params = {
            'day' : str(self.day)
        }
        self.data_list = []

    def get_main(self):
        flag = True
        main_url = 'https://danjuanfunds.com/djapi/fund/{}'.format(self.code)
        try:
            response = requests.get(url=main_url, headers=self.headers)
            data = response.json()['data']
        except:
            print('该基金暂不支持交易。T_T')
        found_date = data['found_date']#成立时间,判断是否大于1年
        str_time = str.split(found_date, '-')
        #now_year = dt.year
        #now_mon = dt.month
        found_year = int(str_time[0])
        found_mon = int(str_time[1])
        found_day = int(str_time[2])
        found_time = dt(found_year, found_mon, found_day)
        delta = dd(days=365)
        if dt.now() - found_time < delta:
            flag = False
        self.data_list.append(data['fd_code'])
        self.data_list.append(data['fd_name'])#前两个数据分别是代码，名称
        fd_name = data['fd_name']
        print('正在爬取写入[' + data['fd_name'] + ']......')
        time.sleep(6)
        return (fd_name, flag)

    def get_growth_data(self):
        growth_url = 'https://danjuanfunds.com/djapi/fund/nav-growth/{}'.format(self.code)
        resp_growth = requests.get(url=growth_url, params=self.params, headers=self.headers)
        growth_data = resp_growth.json()['data']['fund_nav_growth']
        #<时间和涨跌幅的次序不能搞乱！！！>
        for data in growth_data:
            self.data_list.append(data['date'])#bug成立时间不满一年
            if 'gr_per' not in data:
                self.data_list.append(0)
            else:
                self.data_list.append(float(data['gr_per']))

    def all_write(self, sheet_num):
        sheet_name = 'Sheet' + str(sheet_num)
        self.write_to_excel(book_name='danjuan2.xlsx', sheet_name=sheet_name, data_list=self.data_list)

    def write_to_excel(self, book_name, sheet_name, data_list, row=1, col=1):#文件名必须存在
        wb = openpyxl.load_workbook(book_name)
        ws = wb[sheet_name]
        my_row = row
        my_col = col
        for i in range(0, len(data_list), 2):
            ws.cell(my_row, my_col).value = data_list[i]
            ws.cell(my_row, my_col+1).value = data_list[i+1]
            my_row += 1
        #多余填充
        while my_row <= 251:
            ws.cell(my_row, my_col).value = ''
            ws.cell(my_row, my_col+1).value = ''
            my_row += 1
        if my_col < 4:#填充4，5，7，8
            for i in range(1, 12):
                ws.cell(i, 4).value = ''
                ws.cell(i, 5).value = ''
                ws.cell(i, 7).value = ''
                ws.cell(i, 8).value = ''
        if my_col < 7:#填充7, 8
            for i in range(1, 12):
                ws.cell(i, 7).value = ''
                ws.cell(i, 8).value = ''
        wb.save(book_name)

    def stock_hold(self, col):#十大重仓股
        url = 'https://danjuanfunds.com/djapi/fund/detail/' + self.code
        response = requests.get(url=url, headers=self.headers)
        stock_list = response.json()['data']['fund_position']['stock_list']
        name_percent = []#带上代码和名称
        name_percent.append('基金代码：')
        name_percent.append(response.json()['data']['fund_rates']['fd_code'])
        for stock in stock_list:
            name_percent.append(stock['name'])
            name_percent.append(stock['percent'])
            #print((stock['name'] + ' : %.2f%%') % stock['percent'])
        self.write_to_excel('stock_list.xlsx', 'Sheet1', name_percent, col=col)
        print('Done!')


def rename(file_name):
    danjuan = get_path('danjuan2.xlsx')
    new_name = get_path(file_name)
    os.popen('copy {} {}'.format(danjuan, new_name))

def open_file(file_name):
    file_path = get_path(file_name)
    os.startfile(file_path)
    print('Open Ok!')

def get_path(file_name):
    path = os.getcwd()
    new_path = ''
    spl = str.split(path, '\\')
    for i in spl:
        new_path += i
        new_path += '\\'
    new_path += file_name
    return (new_path)

def is_open(file_name):
    temp_file = '~$' + file_name
    if os.path.exists(temp_file):
        xl = Dispatch('Excel.Application')
        wb = xl.Workbooks.open(file_name)
        wb.Close(True)

if __name__ == '__main__':
    code = input('请输入基金代码(q to quit)：')
    count = 1
    day = 360
    name_list = []
    file_name = ''
    while code != 'q' and count < 4:
        fund = Fund(code, day)
        name_flag = fund.get_main()
        name = name_flag[0]
        flag = name_flag[1]
        if not flag:
            print('该基金成立时间不足一年！！！已自动放弃。。。')
            code = input('请输入基金代码(q to quit)：')
            continue
        fund.get_growth_data()
        fund.all_write(count)
        name_list.append(name)
        fund.stock_hold(col = 3*count -2)
        count += 1
        if count < 4:
            code = input('请输入基金代码(q to quit)：')
    for i in range(0, len(name_list)):
        file_name += name_list[i]
    file_name += '.xlsx'
    rename(file_name)
    is_open(file_name)
    open_file('danjuan2.xlsx')#file_name

